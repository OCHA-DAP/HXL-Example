import ckanapi, json, sys, time
from urllib2 import Request, urlopen, URLError, HTTPError
from itertools import islice
from openpyxl import load_workbook
import hxl
from io import BytesIO


DELAY = 1
"""Time delay in seconds between datasets, to give HDX a break."""

CKAN_URL = "https://data.humdata.org"
"""Base URL for the CKAN instance."""

#The ultimate JSON file
indexFile = {}

def populateIndex(uniqueTags,sampleData,i,attributes):
    for tag in uniqueTags:
        if tag not in indexFile:
            indexFile[tag] = {'samples':['sample_'+str(i)],'attributes':{}}
        else:
            if len(indexFile[tag])<5:
                indexFile[tag]['samples'].append('sample_'+str(i))
    print attributes
    for key in attributes:
        for att in attributes[key]:
            if att in indexFile[key]['attributes']:
                indexFile[key]['attributes'][att]+=1
            else:
                indexFile[key]['attributes'][att]=1

    with open('working/sample_'+str(i)+'.json', 'w') as file:
        json.dump(sampleData, file)

    print indexFile

def processHXLData(dataset):
    if len(dataset.values)>2:
        print dataset.headers
        print len(dataset.values)
        print dataset.values[0]
        sample = [dataset.headers,dataset.tags,dataset.values[0],dataset.values[1],dataset.values[2]]
        uniqueTags = []
        atts = {}
        for tag in dataset.tags:
            if tag not in uniqueTags:
                uniqueTags.append(tag)
        for tag in dataset.display_tags:
            tagAtts = tag.split('+')
            if len(tagAtts)>1:
                if tagAtts[0] not in atts:
                    atts[tagAtts[0]]=[]
                for i in range(1, len(tagAtts)):
                    atts[tagAtts[0]].append(tagAtts[i])

        return [uniqueTags,sample,atts]
    return False    

def readCsv(csvLocation):
    try:
        content = urlopen(csvLocation)
    except URLError as e:
        print("CSV Failed to download")
    try:
        print "File downloaded and attempting to read HXL"
        dataset = hxl.data(content).cache()
        output = processHXLData(dataset)
        print "HXL output"
        print output
        return output
    except Exception as e:
        print e
        return False

def readXlsx(fileLocation):
    print "Trying to download XLSX"
    try:
        response = urlopen(fileLocation)
        try:
            wb = load_workbook(BytesIO(response.read()))
        except:
            print "Error reading "+ str(fileLocation)
            return False
        sheet = wb.active
        data={}
    except URLError as e:
        print("XLS Failed to download")
    #try:
    rows_iter = sheet.iter_rows(min_col=1, min_row=1, max_col=sheet.max_column, max_row=sheet.max_row)
    dataset = [[cell.value for cell in row] for row in rows_iter]
    for row in dataset:
        print row
    dataset = hxl.data(dataset)
    output = processHXLData(dataset)
    print "HXL output"
    print output
    return output        
    #except Exception as e:
    #    print e
    #    return False

# find datasets tagged HXL
def find_hxl_datasets(start, rows):
    """Return a page of HXL datasets."""
    return ckan.action.package_search(start=start, rows=rows, fq="tags:hxl")

# Open a connection to HDX
ckan = ckanapi.RemoteCKAN(CKAN_URL)
result_start_pos = 0
result_page_size = 25

result = find_hxl_datasets(result_start_pos, result_page_size)
result_total_count = result["count"]
print(result["count"])
print(result["results"][0]["title"])


packages = result["results"]

# Iterate through all the datasets ("packages") and resources on HDX
i=0
for package in packages:
    # package = ckan.action.package_show(id=package_id)
    print("Package: " + format(package["title"]))

    # for each resource in a package (some packages have multiple csv files for example), print the name, url and format
    for resource in package["resources"]:
        print "---------------------"
        print("  {}".format(resource["name"]))
        print("    {}".format(resource["url"]))
        print resource["format"]

        # if the resource is a csv then print content
        #if resource["format"] == "CSV":
        #    file_data = readCsv(resource["url"])
        #    if(file_data!=False):
        #        populateIndex(file_data[0],file_data[1],i,file_data[2])

        if resource["format"] == "XLSX":
            file_data = readXlsx(resource["url"])
        
        i+=1
    time.sleep(DELAY) # give HDX a short rest
